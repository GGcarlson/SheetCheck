"""Tests for object position validation rule."""

import pytest
from unittest.mock import Mock, patch, MagicMock
from pathlib import Path

from openpyxl import load_workbook

from src.structural.object_pos import ObjectPositionRule
from src.structural.base import ValidationFailure


class TestObjectPositionRule:
    """Test cases for ObjectPositionRule."""

    @pytest.fixture
    def rule_config(self):
        """Create a mock rule configuration."""
        config = Mock()
        config.objects = [
            {
                "name": "RefreshButton",
                "expect_position": {"top": 20, "left": 450, "tolerance": 5},
            }
        ]
        return config

    @pytest.fixture
    def empty_config(self):
        """Create a mock configuration with no objects."""
        config = Mock()
        config.objects = []
        return config

    @pytest.fixture
    def no_objects_config(self):
        """Create a mock configuration without objects attribute."""
        config = Mock()
        # Remove objects attribute to test hasattr check
        if hasattr(config, "objects"):
            delattr(config, "objects")
        return config

    def test_no_objects_configured(self, no_objects_config):
        """Test that rule returns no failures when no objects are configured."""
        rule = ObjectPositionRule("Dashboard", no_objects_config)
        workbook = Mock()
        workbook.sheetnames = ["Dashboard"]

        result = rule.run(workbook)
        assert result == []

    def test_empty_objects_list(self, empty_config):
        """Test that rule returns no failures when objects list is empty."""
        rule = ObjectPositionRule("Dashboard", empty_config)
        workbook = Mock()
        workbook.sheetnames = ["Dashboard"]

        result = rule.run(workbook)
        assert result == []

    def test_sheet_does_not_exist(self, rule_config):
        """Test that rule returns no failures when sheet doesn't exist."""
        rule = ObjectPositionRule("Dashboard", rule_config)
        workbook = Mock()
        workbook.sheetnames = ["OtherSheet"]

        result = rule.run(workbook)
        assert result == []

    @patch("src.structural.object_pos.xw", None)
    def test_xlwings_not_available(self, rule_config):
        """Test that rule returns no failures when xlwings is not available."""
        rule = ObjectPositionRule("Dashboard", rule_config)
        workbook = Mock()
        workbook.sheetnames = ["Dashboard"]

        result = rule.run(workbook)
        assert result == []

    @patch("src.structural.object_pos.xw")
    def test_object_missing(self, mock_xw, rule_config):
        """Test validation failure when expected object is missing."""
        # Mock xlwings app and workbook
        mock_app = MagicMock()
        mock_wb = MagicMock()
        mock_sheet = MagicMock()
        mock_sheet.shapes = []  # No shapes

        mock_xw.App.return_value.__enter__.return_value = mock_app
        mock_app.books.open.return_value = mock_wb
        mock_wb.sheets.__getitem__.return_value = mock_sheet

        rule = ObjectPositionRule("Dashboard", rule_config)
        workbook = Mock()
        workbook.sheetnames = ["Dashboard"]
        workbook.save = Mock()

        result = rule.run(workbook)

        assert len(result) == 1
        failure = result[0]
        assert failure.type == "object_missing"
        assert failure.sheet == "Dashboard"
        assert failure.object == "RefreshButton"
        assert failure.expected == {"top": 20, "left": 450}
        assert failure.found is None
        assert failure.tolerance == 5
        assert "Add shape 'RefreshButton'" in failure.fix_hint

    @patch("src.structural.object_pos.xw")
    def test_object_in_correct_position(self, mock_xw, rule_config):
        """Test no validation failures when object is in correct position."""
        # Mock xlwings app and workbook
        mock_app = MagicMock()
        mock_wb = MagicMock()
        mock_sheet = MagicMock()
        mock_shape = MagicMock()
        mock_shape.name = "RefreshButton"
        mock_shape.top = 15  # 15 pt * 1.333 = 20 px
        mock_shape.left = 337.5  # 337.5 pt * 1.333 = 450 px
        mock_sheet.shapes = [mock_shape]

        mock_xw.App.return_value.__enter__.return_value = mock_app
        mock_app.books.open.return_value = mock_wb
        mock_wb.sheets.__getitem__.return_value = mock_sheet

        rule = ObjectPositionRule("Dashboard", rule_config)
        workbook = Mock()
        workbook.sheetnames = ["Dashboard"]
        workbook.save = Mock()

        result = rule.run(workbook)

        assert result == []

    @patch("src.structural.object_pos.xw")
    def test_object_moved_beyond_tolerance(self, mock_xw, rule_config):
        """Test validation failure when object is moved beyond tolerance."""
        # Mock xlwings app and workbook
        mock_app = MagicMock()
        mock_wb = MagicMock()
        mock_sheet = MagicMock()
        mock_shape = MagicMock()
        mock_shape.name = "RefreshButton"
        mock_shape.top = 15  # 15 pt * 1.333 = 20 px (within tolerance)
        mock_shape.left = (
            350  # 350 pt * 1.333 = 466 px (16 px difference > 5 px tolerance)
        )
        mock_sheet.shapes = [mock_shape]

        mock_xw.App.return_value.__enter__.return_value = mock_app
        mock_app.books.open.return_value = mock_wb
        mock_wb.sheets.__getitem__.return_value = mock_sheet

        rule = ObjectPositionRule("Dashboard", rule_config)
        workbook = Mock()
        workbook.sheetnames = ["Dashboard"]
        workbook.save = Mock()

        result = rule.run(workbook)

        assert len(result) == 1
        failure = result[0]
        assert failure.type == "object_moved"
        assert failure.sheet == "Dashboard"
        assert failure.object == "RefreshButton"
        assert failure.expected == {"top": 20, "left": 450}
        assert failure.found == {"top": 20, "left": 467}
        assert failure.tolerance == 5
        assert "Move 'RefreshButton'" in failure.fix_hint
        assert "left" in failure.fix_hint  # Should suggest moving left

    @patch("src.structural.object_pos.xw")
    def test_object_within_tolerance(self, mock_xw, rule_config):
        """Test no failures when object is slightly moved but within tolerance."""
        # Mock xlwings app and workbook
        mock_app = MagicMock()
        mock_wb = MagicMock()
        mock_sheet = MagicMock()
        mock_shape = MagicMock()
        mock_shape.name = "RefreshButton"
        mock_shape.top = 15  # 15 pt * 1.333 = 20 px
        mock_shape.left = (
            341.25  # 341.25 pt * 1.333 = 455 px (5 px difference = tolerance)
        )
        mock_sheet.shapes = [mock_shape]

        mock_xw.App.return_value.__enter__.return_value = mock_app
        mock_app.books.open.return_value = mock_wb
        mock_wb.sheets.__getitem__.return_value = mock_sheet

        rule = ObjectPositionRule("Dashboard", rule_config)
        workbook = Mock()
        workbook.sheetnames = ["Dashboard"]
        workbook.save = Mock()

        result = rule.run(workbook)

        assert result == []

    @patch("src.structural.object_pos.xw")
    def test_xlwings_exception_handling(self, mock_xw, rule_config):
        """Test that xlwings exceptions are handled gracefully."""
        # Mock xlwings to raise an exception
        mock_xw.App.side_effect = Exception("xlwings failed")

        rule = ObjectPositionRule("Dashboard", rule_config)
        workbook = Mock()
        workbook.sheetnames = ["Dashboard"]
        workbook.save = Mock()

        result = rule.run(workbook)

        # Should return empty list when xlwings fails
        assert result == []

    def test_get_movement_hint_left(self):
        """Test movement hint calculation for leftward movement."""
        rule = ObjectPositionRule("Dashboard", Mock())
        hint = rule._get_movement_hint(
            450, 20, 466, 20
        )  # actual=466, expected=450, need to go left
        assert hint == "left"

    def test_get_movement_hint_right(self):
        """Test movement hint calculation for rightward movement."""
        rule = ObjectPositionRule("Dashboard", Mock())
        hint = rule._get_movement_hint(
            450, 20, 434, 20
        )  # actual=434, expected=450, need to go right
        assert hint == "right"

    def test_get_movement_hint_up(self):
        """Test movement hint calculation for upward movement."""
        rule = ObjectPositionRule("Dashboard", Mock())
        hint = rule._get_movement_hint(
            450, 20, 450, 36
        )  # actual=36, expected=20, need to go up
        assert hint == "up"

    def test_get_movement_hint_down(self):
        """Test movement hint calculation for downward movement."""
        rule = ObjectPositionRule("Dashboard", Mock())
        hint = rule._get_movement_hint(
            450, 20, 450, 4
        )  # actual=4, expected=20, need to go down
        assert hint == "down"

    @patch("src.structural.object_pos.xw")
    def test_multiple_objects_validation(self, mock_xw):
        """Test validation of multiple objects."""
        # Create config with multiple objects
        config = Mock()
        config.objects = [
            {
                "name": "RefreshButton",
                "expect_position": {"top": 20, "left": 450, "tolerance": 5},
            },
            {
                "name": "SaveButton",
                "expect_position": {"top": 50, "left": 500, "tolerance": 10},
            },
        ]

        # Mock xlwings
        mock_app = MagicMock()
        mock_wb = MagicMock()
        mock_sheet = MagicMock()

        # Create mock shapes - one correct, one missing
        mock_shape1 = MagicMock()
        mock_shape1.name = "RefreshButton"
        mock_shape1.top = 15  # Correct position
        mock_shape1.left = 337.5
        mock_sheet.shapes = [mock_shape1]  # SaveButton is missing

        mock_xw.App.return_value.__enter__.return_value = mock_app
        mock_app.books.open.return_value = mock_wb
        mock_wb.sheets.__getitem__.return_value = mock_sheet

        rule = ObjectPositionRule("Dashboard", config)
        workbook = Mock()
        workbook.sheetnames = ["Dashboard"]
        workbook.save = Mock()

        result = rule.run(workbook)

        # Should have one failure for missing SaveButton
        assert len(result) == 1
        assert result[0].type == "object_missing"
        assert result[0].object == "SaveButton"
