"""Tests for SheetExistsRule."""

from openpyxl import load_workbook

from src.structural.sheet_exists import SheetExistsRule
from src.validator.config import load_rules


def _run(rule_yaml: str, wb_path: str):
    """Helper function to run SheetExistsRule with given rule and workbook."""
    cfg = load_rules(rule_yaml)
    wb = load_workbook(wb_path, data_only=True)
    rule = SheetExistsRule("Summary", cfg.sheets["Summary"])
    return rule.run(wb)


def test_success():
    """Test that validation passes when required sheet exists."""
    failures = _run("tests/fixtures/rule_sheet.yaml", "tests/fixtures/good_sheet.xlsx")
    assert failures == []


def test_failure():
    """Test that validation fails when required sheet is missing."""
    failures = _run(
        "tests/fixtures/rule_sheet.yaml", "tests/fixtures/missing_sheet.xlsx"
    )
    assert len(failures) == 1

    failure = failures[0]
    assert failure.type == "sheet_missing"
    assert failure.sheet == "Summary"
    assert failure.fix_hint == "Add a sheet named 'Summary'"


def test_failure_to_dict():
    """Test that validation failure converts properly to dictionary format."""
    failures = _run(
        "tests/fixtures/rule_sheet.yaml", "tests/fixtures/missing_sheet.xlsx"
    )
    assert len(failures) == 1

    failure_dict = failures[0].to_dict()
    expected = {
        "type": "sheet_missing",
        "sheet": "Summary",
        "fix_hint": "Add a sheet named 'Summary'",
        "found": "",
    }
    assert failure_dict == expected


def test_must_exist_false():
    """Test validation passes when must_exist is False and sheet is missing."""
    # Create a temporary config where must_exist is False
    from src.validator.config import SheetCfg

    wb = load_workbook("tests/fixtures/missing_sheet.xlsx", data_only=True)
    sheet_config = SheetCfg(must_exist=False)
    rule = SheetExistsRule("Summary", sheet_config)
    failures = rule.run(wb)
    assert failures == []
