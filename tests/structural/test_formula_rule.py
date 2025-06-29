"""Tests for CellFormulaRule."""

from openpyxl import load_workbook

from src.structural.formula import CellFormulaRule
from src.validator.config import load_rules, SheetCfg


def _run(rule_yaml: str, wb_path: str):
    """Helper function to run CellFormulaRule with given rule and workbook."""
    cfg = load_rules(rule_yaml)
    wb = load_workbook(wb_path, data_only=False)  # Keep formulas
    rule = CellFormulaRule("Summary", cfg.sheets["Summary"])
    return rule.run(wb)


def test_formula_success():
    """Test that validation passes when formula matches."""
    failures = _run(
        "tests/fixtures/rule_formula.yaml", "tests/fixtures/formula_match.xlsx"
    )
    assert failures == []


def test_formula_failure():
    """Test that validation fails when formula is wrong or hard-coded."""
    rule_file = "tests/fixtures/rule_formula.yaml"
    test_file = "tests/fixtures/formula_mismatch.xlsx"
    failures = _run(rule_file, test_file)
    assert len(failures) == 1

    failure = failures[0]
    assert failure.type == "formula_mismatch"
    assert failure.sheet == "Summary"
    assert failure.cell == "B2"
    expected_formula = "=SUM(Data!B2:B10)"
    assert failure.expected == expected_formula
    assert failure.found == "123"
    assert "correct formula" in failure.fix_hint


def test_formula_failure_to_dict():
    """Test that validation failure converts properly to dict format."""
    rule_file = "tests/fixtures/rule_formula.yaml"
    test_file = "tests/fixtures/formula_mismatch.xlsx"
    failures = _run(rule_file, test_file)
    assert len(failures) == 1

    failure_dict = failures[0].to_dict()
    expected_formula = "=SUM(Data!B2:B10)"
    expected = {
        "type": "formula_mismatch",
        "sheet": "Summary",
        "cell": "B2",
        "expected": expected_formula,
        "found": "123",
        "fix_hint": "Replace with correct formula",
    }
    assert failure_dict == expected


def test_case_insensitive_formula_comparison():
    """Test that formula comparison is case-insensitive."""
    from src.validator.config import SheetCfg

    # Create a config with uppercase formula
    sheet_config = SheetCfg(cells={"B2": {"formula": "=sum(data!b2:b10)"}})
    wb = load_workbook("tests/fixtures/formula_match.xlsx", data_only=False)
    rule = CellFormulaRule("Summary", sheet_config)
    failures = rule.run(wb)

    # Should pass because comparison is case-insensitive
    assert failures == []


def test_whitespace_trimming():
    """Test that formula comparison trims whitespace."""
    from src.validator.config import SheetCfg

    # Create a config with extra whitespace
    sheet_config = SheetCfg(cells={"B2": {"formula": "  =SUM(Data!B2:B10)  "}})
    wb = load_workbook("tests/fixtures/formula_match.xlsx", data_only=False)
    rule = CellFormulaRule("Summary", sheet_config)
    failures = rule.run(wb)

    # Should pass because whitespace is trimmed
    assert failures == []


def test_no_cell_config():
    """Test that validation passes when no cell configurations are defined."""
    sheet_config = SheetCfg(must_exist=True)  # No cells config
    wb = load_workbook("tests/fixtures/formula_match.xlsx", data_only=False)
    rule = CellFormulaRule("Summary", sheet_config)
    failures = rule.run(wb)
    assert failures == []


def test_missing_sheet():
    """Test that validation skips when sheet doesn't exist."""
    cfg = load_rules("tests/fixtures/rule_formula.yaml")
    wb = load_workbook("tests/fixtures/missing_sheet.xlsx", data_only=False)
    rule = CellFormulaRule("Summary", cfg.sheets["Summary"])
    failures = rule.run(wb)

    # Should return empty failures (sheet existence handled by other rule)
    assert failures == []


def test_normalize_formula():
    """Test formula normalization function directly."""
    rule = CellFormulaRule("Test", SheetCfg())

    # Test case insensitive
    assert rule._normalize_formula("=SUM(A1:A10)") == "=sum(a1:a10)"

    # Test whitespace trimming
    assert rule._normalize_formula("  =SUM(A1:A10)  ") == "=sum(a1:a10)"

    # Test internal whitespace normalization
    assert rule._normalize_formula("=SUM( A1 : A10 )") == "=sum( a1 : a10 )"

    # Test empty formula
    assert rule._normalize_formula("") == ""
    assert rule._normalize_formula(None) == ""
