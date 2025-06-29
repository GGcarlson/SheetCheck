"""Tests for ConditionalFormattingRule."""

from openpyxl import load_workbook

from src.structural.conditional import ConditionalFormattingRule
from src.validator.config import load_rules, SheetCfg


def _run(rule_yaml: str, wb_path: str):
    """Helper function to run ConditionalFormattingRule with given rule and workbook."""
    cfg = load_rules(rule_yaml)
    wb = load_workbook(wb_path)
    rule = ConditionalFormattingRule("Dashboard", cfg.sheets["Dashboard"])
    return rule.run(wb)


def test_cf_success():
    """Test that validation passes when CF rule exists and matches."""
    failures = _run("tests/fixtures/rule_cf.yaml", "tests/fixtures/cf_ok.xlsx")
    assert failures == []


def test_cf_failure():
    """Test that validation fails when CF rule is missing."""
    failures = _run("tests/fixtures/rule_cf.yaml", "tests/fixtures/cf_missing.xlsx")
    assert len(failures) == 1

    failure = failures[0]
    assert failure.type == "cf_missing"
    assert failure.sheet == "Dashboard"
    assert failure.range == "C2:C50"

    # Check expected structure
    expected = failure.expected
    assert isinstance(expected, dict)
    assert expected["type"] == "colorScale"
    assert expected["colors"] == ["#63BE7B", "#FFEB84", "#F8696B"]

    # Check found is None
    assert failure.found is None
    assert "3-color scale CF to C2:C50" in failure.fix_hint


def test_cf_failure_to_dict():
    """Test that validation failure converts properly to dictionary format."""
    failures = _run("tests/fixtures/rule_cf.yaml", "tests/fixtures/cf_missing.xlsx")
    assert len(failures) == 1

    failure_dict = failures[0].to_dict()
    expected_dict = {
        "type": "cf_missing",
        "sheet": "Dashboard",
        "range": "C2:C50",
        "expected": {"type": "colorScale", "colors": ["#63BE7B", "#FFEB84", "#F8696B"]},
        "found": None,
        "fix_hint": "Add 3-color scale CF to C2:C50",
    }
    assert failure_dict == expected_dict


def test_color_normalization():
    """Test color normalization functions."""
    rule = ConditionalFormattingRule("Test", SheetCfg())

    # Test normalization (removing ARGB prefix, # prefix, uppercase)
    assert rule._normalize_color("#63BE7B") == "63BE7B"
    assert rule._normalize_color("0063BE7B") == "63BE7B"
    assert rule._normalize_color("63be7b") == "63BE7B"
    assert rule._normalize_color("63BE7B") == "63BE7B"

    # Test denormalization (adding # prefix)
    assert rule._denormalize_color("63BE7B") == "#63BE7B"


def test_no_cf_config():
    """Test that validation passes when no CF configurations are defined."""
    sheet_config = SheetCfg(must_exist=True)  # No expect_cf_rules
    wb = load_workbook("tests/fixtures/cf_ok.xlsx")
    rule = ConditionalFormattingRule("Dashboard", sheet_config)
    failures = rule.run(wb)
    assert failures == []


def test_missing_sheet():
    """Test that validation skips when sheet doesn't exist."""
    cfg = load_rules("tests/fixtures/rule_cf.yaml")
    wb = load_workbook("tests/fixtures/missing_sheet.xlsx")
    rule = ConditionalFormattingRule("Dashboard", cfg.sheets["Dashboard"])
    failures = rule.run(wb)

    # Should return empty failures (sheet existence handled by other rule)
    assert failures == []


def test_cf_description():
    """Test CF description generation."""
    rule = ConditionalFormattingRule("Test", SheetCfg())

    # Test color scale description
    assert (
        rule._get_cf_description("colorScale", ["#FF0000", "#00FF00", "#0000FF"])
        == "3-color scale"
    )
    assert (
        rule._get_cf_description("colorScale", ["#FF0000", "#00FF00"])
        == "2-color scale"
    )

    # Test other rule types
    assert rule._get_cf_description("cellValue", []) == "cellValue"


def test_find_matching_cf_rule():
    """Test finding matching CF rule functionality."""
    rule = ConditionalFormattingRule("Test", SheetCfg())
    wb = load_workbook("tests/fixtures/cf_ok.xlsx")
    sheet = wb["Dashboard"]

    # Test finding existing rule
    found = rule._find_matching_cf_rule(
        sheet, "C2:C50", "colorScale", ["#63BE7B", "#FFEB84", "#F8696B"]
    )
    assert found is not None
    assert found["type"] == "colorScale"
    assert found["colors"] == ["#63BE7B", "#FFEB84", "#F8696B"]
    assert found["range"] == "C2:C50"

    # Test not finding non-existent rule
    not_found = rule._find_matching_cf_rule(
        sheet, "D2:D50", "colorScale", ["#FF0000", "#00FF00", "#0000FF"]
    )
    assert not_found is None
