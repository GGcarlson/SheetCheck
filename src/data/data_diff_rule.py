"""Data diff rule for comparing two Excel workbooks cell-by-cell."""

import tempfile
import subprocess
from pathlib import Path
from typing import List, Optional
import pandas as pd
from openpyxl import Workbook, load_workbook

from ..structural.base import Rule, ValidationFailure


class DataDiffRule(Rule):
    """Rule for comparing two Excel workbooks cell-by-cell using xlcompare."""

    def __init__(self, sheet_name: str, sheet_config: Optional[dict] = None) -> None:
        """Initialize data diff rule.
        
        Args:
            sheet_name: Name of the sheet to compare (not used in this context)
            sheet_config: Configuration object (not used for diff rules)
        """
        super().__init__(sheet_name, sheet_config)
        self.id_column = getattr(sheet_config, 'id_column', 'ID') if sheet_config else 'ID'

    def run_diff(self, old_workbook_path: Path, new_workbook_path: Path) -> List[ValidationFailure]:
        """Compare two workbooks and return cell-level differences.
        
        Args:
            old_workbook_path: Path to the old/baseline workbook
            new_workbook_path: Path to the new/current workbook
            
        Returns:
            List of ValidationFailure objects representing cell changes
        """
        failures = []
        
        try:
            # Create temporary directory for xlcompare output
            with tempfile.TemporaryDirectory() as temp_dir:
                temp_path = Path(temp_dir)
                
                # Run xlcompare command with absolute paths
                result = subprocess.run([
                    'xlcompare', 
                    str(old_workbook_path.resolve()), 
                    str(new_workbook_path.resolve())
                ], 
                cwd=temp_path,
                capture_output=True, 
                text=True
                )
                
                if result.returncode != 0:
                    failures.append(ValidationFailure(
                        type="diff_error",
                        message=f"xlcompare failed: {result.stderr}",
                        fix_hint="Check that both workbooks are valid Excel files"
                    ))
                    return failures
                
                # Read the generated diff.xlsx file
                diff_file = temp_path / "diff.xlsx"
                if not diff_file.exists():
                    failures.append(ValidationFailure(
                        type="diff_error", 
                        message="xlcompare did not generate diff.xlsx",
                        fix_hint="Check that workbooks have differences to compare"
                    ))
                    return failures
                
                # Parse the diff file to extract cell changes
                diff_failures = self._parse_diff_file(diff_file, old_workbook_path, new_workbook_path)
                failures.extend(diff_failures)
                
        except Exception as e:
            failures.append(ValidationFailure(
                type="diff_error",
                message=f"Error running diff comparison: {e}",
                fix_hint="Check that xlcompare is installed and workbooks are accessible"
            ))
            
        return failures

    def _parse_diff_file(self, diff_file: Path, old_workbook_path: Path, new_workbook_path: Path) -> List[ValidationFailure]:
        """Parse xlcompare diff.xlsx output to extract cell-level changes.
        
        Args:
            diff_file: Path to the xlcompare diff.xlsx file
            old_workbook_path: Path to old workbook for reference
            new_workbook_path: Path to new workbook for reference
            
        Returns:
            List of ValidationFailure objects for each cell change
        """
        failures = []
        
        try:
            # Read the diff Excel file
            df = pd.read_excel(diff_file)
            
            # Load original workbooks to get precise cell values
            old_wb = load_workbook(old_workbook_path, data_only=True)
            new_wb = load_workbook(new_workbook_path, data_only=True)
            
            # Get the first sheet (xlcompare only processes first sheet)
            old_sheet = old_wb.active
            new_sheet = new_wb.active
            sheet_name = old_sheet.title
            
            # Find rows that have changed
            changed_rows = df[df.get('Changed', 'No') == 'Yes']
            
            for _, row in changed_rows.iterrows():
                # Get the ID to find the row number in original sheets
                id_value = row.get(self.id_column)
                if id_value is None:
                    continue
                    
                # Find the row in original sheets
                old_row_num = self._find_row_by_id(old_sheet, self.id_column, id_value)
                new_row_num = self._find_row_by_id(new_sheet, self.id_column, id_value)
                
                if old_row_num and new_row_num:
                    # Compare each cell in the row
                    cell_failures = self._compare_row_cells(
                        old_sheet, new_sheet, old_row_num, new_row_num, 
                        sheet_name, df.columns
                    )
                    failures.extend(cell_failures)
                elif new_row_num and not old_row_num:
                    # This is a new row
                    failures.append(ValidationFailure(
                        type="row_added",
                        sheet=sheet_name,
                        message=f"New row added with {self.id_column}={id_value}",
                        expected="",
                        found=f"Row {new_row_num}",
                        fix_hint="Row was added to the new workbook"
                    ))
                elif old_row_num and not new_row_num:
                    # This is a deleted row
                    failures.append(ValidationFailure(
                        type="row_deleted", 
                        sheet=sheet_name,
                        message=f"Row deleted with {self.id_column}={id_value}",
                        expected=f"Row {old_row_num}",
                        found="",
                        fix_hint="Row was removed from the new workbook"
                    ))
                        
        except Exception as e:
            failures.append(ValidationFailure(
                type="diff_parse_error",
                message=f"Error parsing diff file: {e}",
                fix_hint="Check that diff.xlsx is a valid Excel file"
            ))
            
        return failures

    def _find_row_by_id(self, sheet, id_column: str, id_value) -> Optional[int]:
        """Find row number by ID column value.
        
        Args:
            sheet: Excel worksheet
            id_column: Name of the ID column
            id_value: Value to search for
            
        Returns:
            Row number (1-based) or None if not found
        """
        try:
            # Find the ID column index
            id_col_idx = None
            for col_idx, cell in enumerate(sheet[1], 1):  # Header row
                if cell.value == id_column:
                    id_col_idx = col_idx
                    break
                    
            if id_col_idx is None:
                return None
                
            # Search for the ID value
            for row_idx in range(2, sheet.max_row + 1):  # Skip header
                cell_value = sheet.cell(row_idx, id_col_idx).value
                if str(cell_value) == str(id_value):
                    return row_idx
                    
        except Exception:
            pass
        
        return None

    def _compare_row_cells(self, old_sheet, new_sheet, old_row_num: int, new_row_num: int, 
                          sheet_name: str, columns: List[str]) -> List[ValidationFailure]:
        """Compare cells between two rows and return failures for differences.
        
        Args:
            old_sheet: Old worksheet
            new_sheet: New worksheet  
            old_row_num: Row number in old sheet
            new_row_num: Row number in new sheet
            sheet_name: Name of the sheet
            columns: List of column names
            
        Returns:
            List of ValidationFailure objects for cell differences
        """
        failures = []
        
        try:
            # Compare each column
            for col_idx, col_name in enumerate(columns, 1):
                if col_name == 'Changed':  # Skip the added 'Changed' column
                    continue
                    
                old_value = old_sheet.cell(old_row_num, col_idx).value
                new_value = new_sheet.cell(new_row_num, col_idx).value
                
                # Convert to string for comparison
                old_str = str(old_value) if old_value is not None else ""
                new_str = str(new_value) if new_value is not None else ""
                
                if old_str != new_str:
                    # Get cell address
                    from openpyxl.utils import get_column_letter
                    cell_address = f"{get_column_letter(col_idx)}{new_row_num}"
                    
                    failures.append(ValidationFailure(
                        type="cell_changed",
                        sheet=sheet_name,
                        cell=cell_address,
                        message=f"Cell {cell_address} changed in column '{col_name}'",
                        expected=old_str,
                        found=new_str,
                        fix_hint=f"Value changed from '{old_str}' to '{new_str}'"
                    ))
                    
        except Exception as e:
            failures.append(ValidationFailure(
                type="cell_compare_error",
                sheet=sheet_name,
                message=f"Error comparing row cells: {e}",
                fix_hint="Check that both sheets have the same structure"
            ))
            
        return failures

    def run(self, workbook: Workbook) -> List[ValidationFailure]:
        """Standard run method for Rule interface (not used for diff).
        
        Args:
            workbook: Excel workbook (not used in diff mode)
            
        Returns:
            Empty list (diff mode uses run_diff instead)
        """
        return []