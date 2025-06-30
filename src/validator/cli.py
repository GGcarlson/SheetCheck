import sys
from pathlib import Path
from typing import List

import click
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from . import __version__
from .config import load_rules, RuleConfigError, RuleConfig
from ..structural import get_structural_rules
from ..structural.base import ValidationFailure
from ..data import DataValidationRule
from ..visual import capture, pixel_diff
from ..reporter import create_reporters


@click.command()
@click.argument("workbook", type=click.Path(exists=True, path_type=Path))
@click.option(
    "--rules",
    type=click.Path(exists=True, path_type=Path),
    default="rules/default.yaml",
    help="Rules YAML file (default: rules/default.yaml)",
)
@click.option(
    "--renderer",
    type=click.Choice(["auto", "com", "html"]),
    default="auto",
    help="Screenshot renderer (default: auto)",
)
@click.option(
    "--report",
    default="json,md",
    help="Comma-separated report formats: json,xml,md,markdown (default: json,md)",
)
@click.option(
    "--update-baseline",
    is_flag=True,
    default=False,
    help="Update baseline files instead of comparing against them",
)
@click.version_option(version=__version__, prog_name="validator")
def main(
    workbook: Path, rules: Path, renderer: str, report: str, update_baseline: bool
) -> None:
    """Validate an Excel workbook."""
    click.echo(f"Validating workbook: {workbook}")
    click.echo(f"Using rules: {rules}")
    click.echo(f"Renderer: {renderer}")
    click.echo(f"Reports: {report}")
    if update_baseline:
        click.echo("Mode: Update baselines")

    # Load and validate rule configuration
    try:
        config = load_rules(rules)
        total_rules = len(config.sheets) + len(config.data_validation_rules)
        click.echo(f"Loaded {total_rules} rule(s)")

        if config.sheets:
            click.echo(f"  - {len(config.sheets)} structural rule(s)")
            for sheet_name, sheet_cfg in config.sheets.items():
                if sheet_cfg.must_exist:
                    click.echo(f"    - {sheet_name}: must exist")

        if config.data_validation_rules:
            click.echo(
                f"  - {len(config.data_validation_rules)} data validation rule(s)"
            )
            for rule_path in config.data_validation_rules:
                click.echo(f"    - {Path(rule_path).name}")

    except RuleConfigError as e:
        click.echo(f"Error loading rules: {e}", err=True)
        raise click.Abort()

    # Load the workbook
    try:
        wb = load_workbook(workbook, data_only=False)
        sheet_list = ", ".join(wb.sheetnames)
        click.echo(f"Loaded workbook with {len(wb.sheetnames)} sheet(s): {sheet_list}")
    except Exception as e:
        click.echo(f"Error loading workbook: {e}", err=True)
        raise click.Abort()

    # Run validation
    structural_failures = run_structural_validation(wb, config)
    data_failures = run_data_validation(wb, config)
    visual_failures = run_visual_validation(
        wb, config, workbook, renderer, update_baseline
    )

    all_failures = structural_failures + data_failures + visual_failures

    # Generate reports using unified reporter system
    try:
        reporters = create_reporters(report, Path("reports"))
        for reporter in reporters:
            output_path = reporter.write_to_file(
                structural_failures, data_failures, visual_failures
            )
            click.echo(f"{reporter.__class__.__name__[:-8]} report written to {output_path}")
    except ValueError as e:
        click.echo(f"Error creating reports: {e}", err=True)

    # Set exit code based on validation results
    if update_baseline:
        # When updating baselines, always exit with 0 if no errors occurred
        if all_failures:
            error_count = len(all_failures)
            click.echo(f"Baseline update failed with {error_count} error(s)")
            if structural_failures:
                click.echo(f"  - {len(structural_failures)} structural failure(s)")
            if data_failures:
                click.echo(f"  - {len(data_failures)} data validation failure(s)")
            if visual_failures:
                click.echo(f"  - {len(visual_failures)} visual validation failure(s)")
            sys.exit(1)
        else:
            click.echo("Baselines updated successfully")
            sys.exit(0)
    else:
        # Normal validation mode
        if all_failures:
            error_count = len(all_failures)
            click.echo(f"Validation failed with {error_count} error(s)")
            if structural_failures:
                click.echo(f"  - {len(structural_failures)} structural failure(s)")
            if data_failures:
                click.echo(f"  - {len(data_failures)} data validation failure(s)")
            if visual_failures:
                click.echo(f"  - {len(visual_failures)} visual validation failure(s)")
            sys.exit(1)
        else:
            click.echo("Validation passed")
            sys.exit(0)


def run_structural_validation(
    workbook: Workbook, config: RuleConfig
) -> List[ValidationFailure]:
    """Run all structural validation rules.

    Args:
        workbook: The loaded Excel workbook
        config: The rule configuration

    Returns:
        List of validation failures
    """
    failures = []
    rule_registry = get_structural_rules()

    for sheet_name, sheet_config in config.sheets.items():
        # Check if sheet existence rule should be applied
        if hasattr(sheet_config, "must_exist") and sheet_config.must_exist:
            rule_class = rule_registry.get("sheet_exists")
            if rule_class:
                rule = rule_class(sheet_name, sheet_config)
                sheet_failures = rule.run(workbook)
                failures.extend(sheet_failures)

        # Check if cell formula validation should be applied
        if hasattr(sheet_config, "cells") and sheet_config.cells:
            rule_class = rule_registry.get("cell_formula")
            if rule_class:
                rule = rule_class(sheet_name, sheet_config)
                sheet_failures = rule.run(workbook)
                failures.extend(sheet_failures)

        # Check if conditional formatting validation should be applied
        if hasattr(sheet_config, "expect_cf_rules") and sheet_config.expect_cf_rules:
            rule_class = rule_registry.get("conditional_formatting")
            if rule_class:
                rule = rule_class(sheet_name, sheet_config)
                sheet_failures = rule.run(workbook)
                failures.extend(sheet_failures)

        # Check if object position validation should be applied
        if hasattr(sheet_config, "objects") and sheet_config.objects:
            rule_class = rule_registry.get("object_position")
            if rule_class:
                rule = rule_class(sheet_name, sheet_config)
                sheet_failures = rule.run(workbook)
                failures.extend(sheet_failures)

    return failures


def run_data_validation(
    workbook: Workbook, config: RuleConfig
) -> List[ValidationFailure]:
    """Run all data validation rules using Great Expectations.

    Args:
        workbook: The loaded Excel workbook
        config: The rule configuration

    Returns:
        List of validation failures from data validation
    """
    failures = []

    for rule_path in config.data_validation_rules:
        try:
            # Create data validation rule and run it
            data_rule = DataValidationRule(rule_path, None)
            rule_failures = data_rule.run(workbook)
            failures.extend(rule_failures)

            click.echo(f"Data validation completed for {Path(rule_path).name}")
            if rule_failures:
                click.echo(f"  - Found {len(rule_failures)} failure(s)")
            else:
                click.echo("  - Passed")

        except Exception as e:
            click.echo(f"Error running data validation for {rule_path}: {e}", err=True)
            failures.append(
                ValidationFailure(
                    type="data_validation_error",
                    message=f"Failed to run data validation: {e}",
                    fix_hint=f"Check rule file: {rule_path}",
                )
            )

    return failures


def run_visual_validation(
    workbook: Workbook,
    config: RuleConfig,
    workbook_path: Path,
    renderer: str,
    update_baseline: bool,
) -> List[ValidationFailure]:
    """Run visual validation by comparing screenshots against baselines.

    Args:
        workbook: The loaded Excel workbook
        config: The rule configuration
        workbook_path: Path to the workbook file
        renderer: Screenshot renderer to use
        update_baseline: Whether to update baselines instead of comparing

    Returns:
        List of visual validation failures
    """
    failures: List[ValidationFailure] = []

    if not capture.is_capture_supported():
        click.echo("Visual validation skipped: No screenshot renderer available")
        return failures

    # Get workbook name for baseline directory
    workbook_name = workbook_path.stem

    for sheet_name in workbook.sheetnames:
        try:
            # Generate screenshot of current sheet
            temp_screenshot_path = Path(f"/tmp/{workbook_name}_{sheet_name}_temp.png")
            temp_screenshot_path.parent.mkdir(parents=True, exist_ok=True)

            capture.capture_sheet_png(
                workbook_path, sheet_name, temp_screenshot_path, renderer=renderer
            )

            if update_baseline:
                # Update mode: copy screenshot to baseline
                baseline_path = get_baseline_path(workbook_name, sheet_name)
                update_baseline_file(temp_screenshot_path, baseline_path)
                click.echo(f"  ✓ Baseline updated for {sheet_name}")

            else:
                # Validation mode: compare against baseline
                baseline_path = get_baseline_path(workbook_name, sheet_name)

                if not baseline_path.exists():
                    failures.append(
                        ValidationFailure(
                            type="visual_baseline_missing",
                            message=f"Baseline image missing for sheet '{sheet_name}'",
                            fix_hint=f"Run with --update-baseline to create baseline: "
                            f"{baseline_path}",
                            sheet=sheet_name,
                        )
                    )
                else:
                    # Compare images using pixel diff
                    diff_ratio = pixel_diff.diff_png(
                        baseline_path, temp_screenshot_path, threshold=0.02
                    )

                    if diff_ratio > 0.01:  # More than 1% difference
                        failures.append(
                            ValidationFailure(
                                type="visual_diff",
                                message=f"Visual difference detected in sheet "
                                f"'{sheet_name}': {diff_ratio:.2%}",
                                fix_hint="Run with --update-baseline to accept changes",
                                sheet=sheet_name,
                            )
                        )
                        click.echo(f"  ✗ Visual diff in {sheet_name}: {diff_ratio:.2%}")
                    else:
                        click.echo(f"  ✓ Visual validation passed for {sheet_name}")

            # Clean up temporary file
            if temp_screenshot_path.exists():
                temp_screenshot_path.unlink()

        except Exception as e:
            click.echo(f"Error in visual validation for {sheet_name}: {e}", err=True)
            failures.append(
                ValidationFailure(
                    type="visual_error",
                    message=f"Visual validation failed for sheet '{sheet_name}': {e}",
                    fix_hint="Check screenshot renderer and file permissions",
                    sheet=sheet_name,
                )
            )

    return failures


def get_baseline_path(workbook_name: str, sheet_name: str) -> Path:
    """Get the baseline image path for a workbook sheet.

    Args:
        workbook_name: Name of the workbook (without extension)
        sheet_name: Name of the sheet

    Returns:
        Path to the baseline PNG file
    """
    return Path("baselines") / "sheets" / workbook_name / f"{sheet_name}.png"


def update_baseline_file(source_path: Path, baseline_path: Path) -> None:
    """Update a baseline file by copying from source.

    Args:
        source_path: Path to the new screenshot/capture
        baseline_path: Path to the baseline file to update
    """
    import shutil

    # Ensure baseline directory exists
    baseline_path.parent.mkdir(parents=True, exist_ok=True)

    # Copy the new file to baseline location
    shutil.copy2(source_path, baseline_path)




if __name__ == "__main__":
    main()
